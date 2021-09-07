#! /usr/bin/env python
# -*- coding: utf-8 -*-
import requests
import json
from openpyxl import load_workbook

clkDataWbAddr = '/Users/cc/Downloads/PCIM 数据20210906_181327.xlsx' # 数据表
idWbAddr = '/Users/cc/Downloads/IM 数据tableExport .xlsx' # ID 表
CLKCOLUMN = 1 # 数据表中，ID 所在列
IDCOLUMN = 3 # ID 表中，ID 所在列

# 读取数据表
clkWb = load_workbook(clkDataWbAddr) 
clkWs = clkWb.worksheets[2] # 点击数据表的第几个 sheet
idWb = load_workbook(idWbAddr)
idWs = idWb.worksheets[0] # ID 表的第几个 sheet

# 遍历 id 表
for index in range(clkWs.max_row):
    if(index<=2): # 跳过点击数据表的表头
        continue
    v = clkWs.cell(row = index,column = CLKCOLUMN).value # ID 所在列

    for j in range(idWs.max_row):

        if(j<2): # 跳过 ID 表的表头
            continue
        if(v == idWs.cell(row = j,column = IDCOLUMN).value): # 如果在 ID 表中匹配到 v（iD），则提取 id 表中的值填入数据表中
            clkWs.cell(row = index,column = 3).value = idWs.cell(row = j,column = 1).value # ID 数据表某列的数据，例如分类名称
            clkWs.cell(row = index,column = 4).value = idWs.cell(row = j,column = 2).value # ID 数据表某列的数据，例如事件名称
            
            # 匹配到结果后跳出此处循环
            break 

clkWb.save('Py后的处理数据.xlsx')